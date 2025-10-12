#Calculo para la determinación del WACC

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime 

# Obtener el directorio actual donde está el script de Python
base_dir = os.getcwd()

# Ruta completa al archivo Excel en la subcarpeta Flujo/input
Entrada = os.path.join(base_dir, 'Flujo', 'input', 'Datos-wacc.xlsx')

# Cargar los datos desde el archivo Excel
df = pd.read_excel(Entrada, engine='openpyxl')

# Verifica que las columnas que tienes en el archivo son adecuadas para la variación porcentual
print(df.head())  # Esto te ayudará a ver cómo están estructurados los datos

# Asegúrate de que las fechas estén en formato datetime
df['Fecha'] = pd.to_datetime(df['Fecha'])  # Asegúrate de que la columna 'Fecha' esté en formato datetime

# Calculamos la variación porcentual de las variables
df_variacion = df.copy()  # Hacemos una copia del DataFrame para mantener los datos originales intactos

# Calculamos la variación porcentual solo para las columnas numéricas
for col in df.columns:
    if col != 'Fecha':  # No calculamos la variación porcentual para la columna 'Fecha'
        df_variacion[col] = df[col].pct_change()  # pct_change calcula la variación porcentual en formato decimal

# Ver el DataFrame con la variación porcentual
print(df_variacion.head())

# Ahora, exportamos el DataFrame a un nuevo archivo Excel
fecha_hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # Obtener la fecha y hora
Resultado_variacion = os.path.join(base_dir, 'Flujo', 'output', f'variacion_porcentual_wacc_{fecha_hora}.xlsx')

# Exportar a un nuevo archivo Excel con las variaciones porcentuales (sin multiplicar por 100)
df_variacion.to_excel(Resultado_variacion, index=False, engine='openpyxl')

# Cargar el archivo Excel exportado
wb = load_workbook(Resultado_variacion)
ws = wb.active

# Crear un estilo para el formato porcentaje
percentage_style = NamedStyle(name="percentage_style", number_format="0.00%")

# Crear un estilo para el formato de fecha corta
date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")  # Puedes ajustar el formato de fecha si es necesario

# Aplicar el estilo de porcentaje a las celdas que contienen variaciones porcentuales (todas excepto la columna 'Fecha')
for col in range(2, len(df.columns) + 1):  # Comienza desde la columna 2 para omitir 'Fecha'
    for row in range(2, len(df) + 2):  # Comienza desde la fila 2 para omitir encabezados
        cell = ws.cell(row=row, column=col)
        cell.style = percentage_style

# Aplicar el formato de fecha corta a la columna 'Fecha'
for row in range(2, len(df) + 2):  # Comienza desde la fila 2 para omitir encabezados
    cell = ws.cell(row=row, column=1)  # Columna 1 es la columna 'Fecha'
    cell.style = date_style

# Guardar el archivo con los nuevos formatos
wb.save(Resultado_variacion)

# Confirmación
print(f'Archivo exportado a: {Resultado_variacion}')
